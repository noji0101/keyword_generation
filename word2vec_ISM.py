import xlwings as xw
import datetime
import MeCab

import warnings
# 警告を無視
warnings.filterwarnings(action='ignore', category=UserWarning, module='gensim') 
warnings.filterwarnings(action='ignore', category=FutureWarning, module='gensim')
warnings.filterwarnings(action='ignore', category=DeprecationWarning, module='gensim')

from gensim.models import word2vec
from gensim.models import KeyedVectors
from gensim.models.word2vec import LineSentence
# import better_exceptions
import re

import pickle
import random
import math

import openpyxl
from openpyxl.styles.fills import PatternFill


def main():
    # 日時を記録してあるファイルを開き，日時または回数を取得
    wb = openpyxl.load_workbook('/Users/ryogo/Desktop/research/word2vec_ISM/date_records.xlsx')
    sheet_record = wb['Sheet1']
    opendate = sheet_record.cell(row=1, column=2).value
    openiter = sheet_record.cell(row=2, column=2).value
    opensubject = sheet_record.cell(row=3, column=2).value

    print(opendate,openiter,opensubject)

    # さっき取得した日時または回数を利用して最新のエクセルファイルを開く
    # data_only=Trueはエクセルの数式ではなく値を読み込むという意味
    wbi = openpyxl.load_workbook(f'/Users/ryogo/Desktop/research/word2vec_ISM/system_by_ISM_{opensubject}_{opendate}_{openiter}.xlsm',
                                 keep_vba=True, data_only=True)
    # wb = openpyxl.load_workbook('system_by_ISM.xlsm', keep_vba=True)
    sheet_input = wbi['Sheet3']
    sheet_result = wbi['実行結果']

    # エクセルからの入力された価値、意味、状態、属性を変数に代入
    youso = sheet_input.cell(row=3, column=1).value  # 設計対象

    value_elem = sheet_input.cell(row=6, column=2).value
    meaning_elem = sheet_input.cell(row=7, column=2).value
    state_elem = sheet_input.cell(row=8, column=2).value
    attr_elem = sheet_input.cell(row=9, column=2).value

    value_elem_2 = sheet_input.cell(row=6, column=3).value
    mean_elem_2 = sheet_input.cell(row=7, column=3).value
    state_elem_2 = sheet_input.cell(row=8, column=3).value
    attr_elem_2 = sheet_input.cell(row=9, column=3).value

    # 学習データをオープン
    with open('/Users/ryogo/Desktop/research/word2vec_ISM/word2vec.binaryfile', 'rb') as web:
        model = pickle.load(web)

    # word2vecに単語が存在するか入れてみてチェック なければエラーがでて中止
    testword2vec1 = model.wv.most_similar(positive=[value_elem])
    testword2vec1 = model.wv.most_similar(positive=[meaning_elem])
    testword2vec1 = model.wv.most_similar(positive=[state_elem])
    testword2vec1 = model.wv.most_similar(positive=[attr_elem])

    # 日時記録用.elsxに入力価値、意味、状態、属性をISM作成回数に応じた行に書き込む
    if value_elem_2 != None:
        testword2vec1 = model.wv.most_similar(positive=[value_elem_2])
        sheet_record.cell(row=openiter + 2, column=14, value=value_elem + '+' + value_elem_2) 
    else:
        sheet_record.cell(row=openiter + 2, column=14, value=value_elem) 
    if mean_elem_2 != None:
        testword2vec1 = model.wv.most_similar(positive=[mean_elem_2])
        sheet_record.cell(row=openiter + 2, column=15, value=meaning_elem + '+' + mean_elem_2) 
    else:
        sheet_record.cell(row=openiter + 2, column=15, value=meaning_elem) 
    if state_elem_2 != None:
        testword2vec1 = model.wv.most_similar(positive=[state_elem_2])
        sheet_record.cell(row=openiter + 2, column=16, value=state_elem + '+' + state_elem_2) 
    else:
        sheet_record.cell(row=openiter + 2, column=16, value=state_elem) 
    if attr_elem_2 != None:
        testword2vec1 = model.wv.most_similar(positive=[attr_elem_2])
        sheet_record.cell(row=openiter + 2, column=17, value=attr_elem + '+' + attr_elem_2) 
    else:
        sheet_record.cell(row=openiter + 2, column=17, value=attr_elem) 

    # ????
    追加学習数 = sheet_result.cell(row=1, column=10).value
    新たに選択された要素数 = sheet_result.cell(row=2, column=10).value
    新たに選択された連関数 = sheet_result.cell(row=3, column=10).value
    良い要素数 = sheet_result.cell(row=4, column=10).value
    良い連関数 = sheet_result.cell(row=5, column=10).value
    発想された要素数 = sheet_result.cell(row=6, column=10).value
    発想された連関数 = sheet_result.cell(row=7, column=10).value

    # 記録用シートに評価値の記録
    if openiter > 0:
        sheet_record.cell(row=openiter + 1, column=4, value=str(openiter) + '回目')
        sheet_record.cell(row=openiter + 1, column=5, value=追加学習数)
        sheet_record.cell(row=openiter + 1, column=6, value=新たに選択された要素数)
        sheet_record.cell(row=openiter + 1, column=7, value=新たに選択された連関数)
        sheet_record.cell(row=openiter + 1, column=8, value=良い要素数)
        sheet_record.cell(row=openiter + 1, column=9, value=良い連関数)
        sheet_record.cell(row=openiter + 1, column=10, value=発想された要素数)
        sheet_record.cell(row=openiter + 1, column=11, value=発想された連関数)
        sheet_record.cell(row=openiter + 1, column=12, value=良い要素数 / 発想された要素数) # 良い要素割合
        sheet_record.cell(row=openiter + 1, column=13, value=良い連関数 / 発想された連関数) # 良い連関割合

    wb.save('/Users/ryogo/Desktop/research/word2vec_ISM/date_records.xlsx')

    # 各種パラメータ
    topword = 20  # 一度に提示する共起単語の数を指定
    ISMsize = sheet_input.cell(row=15, column=1).value  # ISMを構成する要素の数を指定
    nlearn = sheet_input.cell(row=21, column=1).value  # 追加学習回数

   
    if openiter > 0:
         # ISMから学習させる文の読み込み + ????
        for i in range(1, 1000):
            file = open(f'/Users/ryogo/Desktop/research/word2vec_ISM/additional_traindata/{opendate}_{openiter}回目の追加学習データ_{opensubject}.txt', 'a')  # 追加書き込みモードでオープン
            fileall = open(f'/Users/ryogo/Desktop/research/word2vec_ISM/additional_traindata/学習データ全部_{opensubject}.txt', 'a')  # 追加書き込みモードでオープン
            if sheet_result.cell(row=i, column=1).value == None:
                break
            gakushu = sheet_result.cell(row=i, column=1).value
            string = gakushu + "\n"  # \nは改行
            file.write(string)
            fileall.write(string)

        # 形態素解析で分かち書き
        class Wakati:
            def __init__(self, file_dir, dic_dir=None, user_dir=None, hinshis=["動詞", "形容詞", "形容動詞", "助動詞"]):

                # if dic_dir is not None and user_dir is not None:
                #     self.tagger = MeCab.Tagger("mecabrc -d {} -u {}".format(dic_dir, user_dir))
                #     # self.tagger = MeCab.Tagger("C:/Users/fut-sal/Desktop/wikiextractor/AA/mecab-ipadic-neologd")
                # elif dic_dir is not None:
                #     self.tagger = MeCab.Tagger("mecabrc -d {}".format(dic_dir))
                #     # self.tagger = MeCab.Tagger("C:/Users/fut-sal/Desktop/wikiextractor/AA/mecab-ipadic-neologd")
                # else:
                #     self.tagger = MeCab.Tagger("mecabrc")
                #     # self.tagger = MeCab.Tagger("C:/Users/fut-sal/Desktop/wikiextractor/AA/mecab-ipadic-neologd")
                self.tagger = MeCab.Tagger("/Users/ryogo/Desktop/research/word2vec_ISM/mecab-ipadic-neologd")
                self.f = open(file_dir, 'r')  ###########,encoding="utf-8" 追加
                self.hinshis = hinshis
                self.splited_text = None
                self.out_dir = None

            def wakati(self):
                line = self.f.readline()
                splited_text = []
                while line:
                    node = self.tagger.parseToNode(line).next
                    splited_line = []
                    while node.surface:
                        word = node.surface
                        feature = node.feature.split(',')
                        hinshi = feature[0]
                        kata = feature[5]
                        genkei = feature[6]
                        if hinshi in self.hinshis:
                            if kata != "基本形":
                                word = genkei
                        splited_line.append(word)
                        node = node.next
                    splited_text.append(splited_line)
                    line = self.f.readline()
                self.splited_text = splited_text
                self.f.close()

            def output(self, out_dir):
                assert self.splited_text is not None
                if self.out_dir is None:
                    self.out_dir = out_dir
                self.fout = open(self.out_dir, 'w', encoding="utf-8")  ###########,encoding="utf-8" 追加
                for line in self.splited_text:
                    self.fout.write(" ".join(line) + " ")
                self.fout.close()

        w = Wakati(f'/Users/ryogo/Desktop/research/word2vec_ISM/additional_traindata/{opendate}_{openiter}回目の追加学習データ_{opensubject}.txt')
        w.wakati()
        w.output(f'/Users/ryogo/Desktop/research/word2vec_ISM/additional_traindata/{opendate}_{openiter}回目の追加学習データのわかちがき_{opensubject}.txt')

        # 追加学習させる
        lee_data = LineSentence(f'/Users/ryogo/Desktop/research/word2vec_ISM/additional_traindata/{opendate}_{openiter}回目の追加学習データのわかちがき_{opensubject}.txt')
        model.build_vocab(lee_data, update=True)
        model.train(lee_data, total_examples=model.corpus_count)  # 学習する部分
        # epochs=nlearnはバージョンの違いで削除

        # 学習データを更新
        with open('/Users/ryogo/Desktop/research/word2vec_ISM/word2vec.binaryfile', 'wb') as web:  # 保存する方
            pickle.dump(model, web)

        print('再学習完了')

    # データ保存の配列
    checklist = [value_elem, meaning_elem, state_elem, attr_elem]  # 重複検索に使うつもり

    # これまで発想された各空間要素リスト ??? ひとつなのでは？
    value_list = [value_elem]
    mean_list = [meaning_elem]
    state_list = [state_elem]
    attr_list = [attr_elem]

    # 今まで連関したリスト
    renkan_vm = [[value_elem, meaning_elem]]
    renkan_ms = [[meaning_elem, state_elem]]
    renkan_sa = [[state_elem, attr_elem]]

    renkan_vv = []
    renkan_mm = []
    renkan_ss = []
    renkan_aa = []

    # value_list = []

    # 与えられた属性要素から属性要素を発想
    # 価値、意味、状態、属性を引数にもち、それに応じたリストを作る。
    def hassou(x):
        if x == '価値':
             main_list = value_list
             renkan = renkan_vv
             input_2 = value_elem_2
        if x == '意味':
             main_list = mean_list
             renkan = renkan_mm
             input_2 = mean_elem_2
        if x == '状態':
             main_list = state_list
             renkan = renkan_ss
             input_2 = state_elem_2
        if x == '属性':
             main_list = attr_list
             renkan = renkan_aa
             input_2 = attr_elem_2

        for i in range(1):# ????
            wordlist = []
            # ランダムに選択する
            n_random = math.floor(random.uniform(0, len(main_list)))
            n_random = 0  # ここを消せばランダムなる

            if input_2 == None:
                入力単語集合 = [main_list[n_random]]
            else:
                入力単語集合 = [main_list[n_random], input_2]
            
            # 最も類似度が高い単語を抽出
            kyouki = model.most_similar(positive=入力単語集合, topn=topword)

            # 記録リストの調整
            for result in kyouki:
                wordlist.append(result[0])  # 結果からワードだけを取り出している，値だけでも可能
            print(x + '空間内での共起') # xは価値、意味、状態、属性のどれか
            print(main_list[n_random])
            print(wordlist, end='')

            
            # topword:一度に提示する共起単語の数
            for j in range(topword):
                # 重複チェック
                if not wordlist[j] in value_list and \
                not wordlist[j] in mean_list and \
                not wordlist[j] in state_list and \
                not wordlist[j] in attr_list:
                    # if not wordlist[j] in usedv and not wordlist[j] in usedm and not wordlist[j] in useds and not wordlist[j] in useda:
                    main_list.append(wordlist[j])  # 今まで共起した単語と重複しないj+1番めの順位の共起語を持ってくる
                    renkan.append([main_list[n_random], wordlist[j]])
                    break

        print('採用された単語：' + wordlist[j]) # ???

    # 与えられた属性要素，状態要素から新たな状態要素を発想 seni=遷移
    def seni(x):
        if x == '属性から状態':
            beforelist = attr_list 
            afterlist = state_list
            renkan = renkan_sa
            input_2a = attr_elem_2
            input_2b = state_elem_2
        if x == '状態から意味':
            beforelist = state_list 
            afterlist = mean_list
            renkan = renkan_ms
            input_2a = state_elem_2
            input_2b = mean_elem_2
        if x == '意味から価値':
            beforelist = mean_list 
            afterlist = value_list
            renkan = renkan_vm
            input_2a = mean_elem_2
            input_2b = value_elem_2
        if x == '価値から意味':
            beforelist = value_list 
            afterlist = mean_list
            renkan = renkan_vm
            input_2a = value_elem_2
            input_2b = mean_elem_2
        if x == '意味から状態':
            beforelist = mean_list 
            afterlist = state_list
            renkan = renkan_ms
            input_2a = mean_elem_2
            input_2b = state_elem_2
        if x == '状態から属性':
            beforelist = state_list 
            afterlist = attr_list
            renkan = renkan_sa
            input_2a = state_elem_2
            input_2b = attr_elem_2

        r_random = math.floor(random.uniform(0, len(renkan)))
        bn_random = math.floor(random.uniform(1, len(beforelist)))

        # renkan[r_random][0] #価値1
        # 選択した連関と選ばれた単語が等しい場合だめなので、異なるまでランダムを回す
        if x == '属性から状態' or x == '状態から意味' or x == '意味から価値':  # 要するに下から上の場合
            while renkan[r_random][1] == beforelist[bn_random]:  
                r_random = math.floor(random.uniform(0, len(renkan)))
                bn_random = math.floor(random.uniform(1, len(beforelist)))
        else:  # 要するに上から下の場合
            while renkan[r_random][0] == beforelist[bn_random]:
                r_random = math.floor(random.uniform(0, len(renkan)))
                bn_random = math.floor(random.uniform(1, len(beforelist)))

        # 既存1-既存2+発生源1=新要素
        r_random = 0  # ここを消せばランダムな連関が選択される 0にすると自分で入力した要素の連関

        if x == '属性から状態' or x == '状態から意味' or x == '意味から価値':  
            # 要するに下から上の場合
            if input_2b == None:
                連関の前側の単語集合 = [renkan[r_random][0], beforelist[bn_random]]
            else:
                連関の前側の単語集合 = [renkan[r_random][0], input_2b, beforelist[bn_random]]

            if input_2a == None:
                連関の後ろ側の単語集合 = [renkan[r_random][1]]
            else:
                連関の後ろ側の単語集合 = [renkan[r_random][1], input_2a]

            kyouki = model.most_similar(positive=連関の前側の単語集合,
                                           negative=連関の後ろ側の単語集合,
                                           topn=topword)

        else:  # 要するに上から下の場合

            if input_2a == None:
                連関の前側の単語集合 = [renkan[r_random][1], beforelist[bn_random]]
            else:
                連関の前側の単語集合 = [renkan[r_random][1], input_2a, beforelist[bn_random]]

            if input_2b == None:
                連関の後ろ側の単語集合 = [renkan[r_random][0]]
            else:
                連関の後ろ側の単語集合 = [renkan[r_random][0], input_2b]

            kyouki = model.most_similar(positive=連関の前側の単語集合,
                                           negative=連関の後ろ側の単語集合,
                                           topn=topword)

        # 共起を表示する
        wordlist = []
        for result in kyouki:
            wordlist.append(result[0])
        print(x + 'を発想')
        if x == '属性から状態' or x == '状態から意味' or x == '意味から価値':  # 要するに下から上の場合
            print(renkan[r_random][0] + ' - ' + renkan[r_random][1] + ' + ' + beforelist[bn_random])
        else:
            print(renkan[r_random][1] + ' - ' + renkan[r_random][0] + ' + ' + beforelist[bn_random])
        print(wordlist, end='')

        # リストの調整
        # afuse.append(afterlist[0])
        # afterlist.remove(afterlist[0])
        for j in range(topword):
            # 重複チェック
            if not wordlist[j] in value_list \
            and not wordlist[j] in mean_list \
            and not wordlist[j] in state_list \
            and not wordlist[j] in attr_list:
                # if not wordlist[j] in usedv and not wordlist[j] in usedm and not wordlist[j] in useds and not wordlist[j] in useda:
                afterlist.append(wordlist[j])  # 重複した場合j番めの順位の共起語を持ってくる
                break
        print('採用された単語：' + wordlist[j])
        # 連関の記録
        if x == '属性から状態' or x == '状態から意味' or x == '意味から価値':
            renkan.append([wordlist[j], beforelist[bn_random]])
        else:
            renkan.append([beforelist[bn_random], wordlist[j]])

    # ランダムスタート空間 = random.uniform(0, 1)
    # if ランダムスタート空間 < 1 / 4:
    #     現在の空間 = '価値'
    # elif ランダムスタート空間 < 2 / 4:
    #     現在の空間 = '意味'
    # elif ランダムスタート空間 < 3 / 4:
    #     現在の空間 = '状態'
    # elif ランダムスタート空間 <= 1:
    #     現在の空間 = '属性'

    # 現在の空間 = '価値'  # スタートする空間 これもランダムにするか

    hassou('価値')
    hassou('意味')
    hassou('状態')
    hassou('属性')

# 確率的にどんな発想や遷移を行うかを決める
# ISMsize:ISMを構成する要素の数
    for i in range(ISMsize - 8):
        n_randomdom = random.uniform(0, 1)
        if n_randomdom < 0.3: # 確率0.3で
            n_randomdom = random.uniform(0, 1)
            if n_randomdom < 0.6:
                hassou('属性')
            elif n_randomdom <= 1:
                seni('属性から状態')
                # 現在の空間 = '状態'
        elif n_randomdom < 0.5:  # 確率0.2で
            n_randomdom = random.uniform(0, 1)
            if n_randomdom < 3 / 10:
                hassou('状態')
            elif n_randomdom < 6 / 10:
                seni('状態から意味')
                # 現在の空間 = '意味'
            elif n_randomdom <= 1:
                seni('状態から属性')
                # 現在の空間 = '属性'
        elif n_randomdom < 0.7: # 確率0.2で
            n_randomdom = random.uniform(0, 1)
            if n_randomdom < 3/ 10:
                hassou('意味')
            elif n_randomdom < 7 / 10:
                seni('意味から価値')
                # 現在の空間 = '価値'
            elif n_randomdom <= 1:
                seni('意味から状態')
                # 現在の空間 = '状態'
        elif n_randomdom <= 1: # 確率0.3で
            n_randomdom = random.uniform(0, 1)
            if n_randomdom < 0.6:
                hassou('価値')
            elif n_randomdom <= 1:
                seni('価値から意味')
                # 現在の空間 = '意味'

    #######################################エクセルに出力

    # さっき取得した日時または回数を利用して最新のエクセルファイルを開く
    # エクセルから呼び出さないとエラーになる
    wb = xw.Book.caller()
    wb.sheets['ISM実行'].range('d14:zz999').value = None

    # エクセルの開始行，列
    offrow = 14
    offcol = 5

    def excel(x):
        # global strow
        # global stcol
        if x == '価値':
            main_list = value_list
            renkan = renkan_vv
            strow = 14
            stcol = 5
            word2 = value_elem_2
        if x == '意味':
            main_list = mean_list
            renkan = renkan_mm
            strow = 14 + len(value_list)
            stcol = 5 + len(value_list)
            word2 = mean_elem_2
        if x == '状態':
            main_list = state_list
            renkan = renkan_ss
            strow = 14 + len(value_list) + len(mean_list)
            stcol = 5 + len(value_list) + len(mean_list)
            word2 = state_elem_2
        if x == '属性':
            main_list = attr_list
            renkan = renkan_aa
            strow = 14 + len(value_list) + len(mean_list) + len(state_list)
            stcol = 5 + len(value_list) + len(mean_list) + len(state_list)
            word2 = attr_elem_2
        for i in range(len(main_list)):
            # 要素名を入力
            if i == 0 and word2 != None:
                wb.sheets['ISM実行'].range((offrow, stcol + 1 + i)).value = main_list[i] + word2
                wb.sheets['ISM実行'].range((strow + 1 + i, offcol)).value = main_list[i] + word2
            else:
                wb.sheets['ISM実行'].range((offrow, stcol + 1 + i)).value = main_list[i]
                wb.sheets['ISM実行'].range((strow + 1 + i, offcol)).value = main_list[i]
            wb.sheets['ISM実行'].range((strow + 1 + i, stcol + 1 + i)).value = 1
            wb.sheets['ISM実行'].range((strow + 1 + i, offcol - 1)).value = x + '要素'

        # 要素を入力（DSM部分)
        for i in range(0, len(renkan)):
            rown = main_list.index(renkan[i][0])
            coln = main_list.index(renkan[i][1])
            wb.sheets['ISM実行'].range((strow + 1 + rown, stcol + 1 + coln)).value = 1
            # sheet_record.cell(row=strow + 1 + rown, column=stcol + 1 + coln, value=1)

        # strow = strow + len(main_list)
        # stcol = stcol + len(main_list)

    # 価値意味，意味状態，状態属性のDMMを入力
    for i in range(len(renkan_vm)):
        rown = offrow + 1 + value_list.index(renkan_vm[i][0])
        coln = offcol + 1 + len(value_list) + mean_list.index(renkan_vm[i][1])
        wb.sheets['ISM実行'].range((rown, coln)).value = 1
        # sheet_record.cell(row=rown, column=coln, value=1)

    for i in range(len(renkan_ms)):
        rown = offrow + 1 + len(value_list) + mean_list.index(renkan_ms[i][0])
        coln = offcol + 1 + len(value_list) + len(mean_list) + state_list.index(renkan_ms[i][1])
        wb.sheets['ISM実行'].range((rown, coln)).value = 1
        # sheet_record.cell(row=rown, column=coln, value=1)

    for i in range(len(renkan_sa)):
        rown = offrow + 1 + len(value_list) + len(mean_list) + state_list.index(renkan_sa[i][0])
        coln = offcol + 1 + len(value_list) + len(mean_list) + len(state_list) + attr_list.index(renkan_sa[i][1])
        wb.sheets['ISM実行'].range((rown, coln)).value = 1
        # sheet_record.cell(row=rown, column=coln, value=1)

    excel('価値')
    excel('意味')
    excel('状態')
    excel('属性')

    # 設計対象を入力
    wb.sheets['ISM実行'].range((13, 5)).value = youso
    # sheet_record.cell(row=13, column=5, value=youso)

    dt_now = datetime.datetime.now()
    dt_now = str(dt_now.replace(microsecond=0))
    dt_now = dt_now.replace(':', '-')

    wb.sheets['Sheet3'].range('A1').value = dt_now
    wb.sheets['Sheet3'].range('B1').value = openiter + 1

    # 最後にアクティブシートを選択する
    ws = wb.sheets['Sheet3']
    ws.activate()

    # wb.save('ISM・行列並び替え.xlsm') 
    wb.save(f'/Users/ryogo/Desktop/research/word2vec_ISM/system_by_ISM_{opensubject}_{dt_now}_{openiter + 1}.xlsm')

    # 日時または回数を更新
    wb = openpyxl.load_workbook('/Users/ryogo/Desktop/research/word2vec_ISM/date_records.xlsx')
    sheet_record = wb['Sheet1']
    sheet_record.cell(row=1, column=2, value=dt_now)
    sheet_record.cell(row=2, column=2, value=openiter + 1)
    wb.save('/Users/ryogo/Desktop/research/word2vec_ISM/date_records.xlsx')

zenbu()
